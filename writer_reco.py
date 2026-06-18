"""
writer_reco.py — AIS reconciliation deliverable (xlsx, Arial 11).

Sheets:
  - "Reco Summary"  : counts + sale-value totals (CG vs AIS) and the net delta
  - "Mismatched"    : same security, sale value differs — the chase list
  - "Only in CG"    : in the broker/CG file, not in AIS
  - "Only in AIS"   : in AIS, not in the broker/CG file
  - "Matched"       : agrees within tolerance (kept for the full audit trail)

Mirrors writer_summary's look (Arial 11, accounting format, thin borders).
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
SZ = 11
ACCT = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
GREY = PatternFill("solid", fgColor="EFEFEF")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREEN = PatternFill("solid", fgColor="E6F4EA")
RED = PatternFill("solid", fgColor="FBE6E6")


def _f(cell, bold=False, size=SZ, fill=None, align=None):
    cell.font = Font(name=ARIAL, size=size, bold=bold)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")


def _amt(cell, val, bold=False, fill=None):
    cell.value = val
    cell.number_format = ACCT
    _f(cell, bold=bold, fill=fill)


def write_reco(result, path, cg_label="CG file", ais_label="AIS"):
    wb = Workbook()
    _summary_sheet(wb, result, cg_label, ais_label)
    _mismatch_sheet(wb, result, cg_label, ais_label)
    _side_sheet(wb, "Only in CG", result.only_cg, RED)
    _side_sheet(wb, "Only in AIS", result.only_ais, RED)
    _matched_sheet(wb, result, cg_label, ais_label)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(path)
    return path


def _summary_sheet(wb, result, cg_label, ais_label):
    ws = wb.create_sheet("Reco Summary")
    ws["A1"] = "AIS Reconciliation"; _f(ws["A1"], bold=True, size=12)
    ws["A2"] = f"{cg_label}  ⟷  {ais_label}   |   computer-prepared, preparer to verify"
    _f(ws["A2"], size=9)
    c = result.counts(); t = result.totals()
    rows = [
        ("", "Securities", "Sale value"),
        ("Matched", c["matched"], None),
        ("Mismatched", c["mismatched"], None),
        ("Only in CG", c["only_cg"], t["cg"]),
        ("Only in AIS", c["only_ais"], t["ais"]),
    ]
    r = 4
    for label, n, val in rows:
        ws.cell(r, 1, label); _f(ws.cell(r, 1), bold=(r == 4), fill=GREY if r == 4 else None)
        ws.cell(r, 2, n); _f(ws.cell(r, 2), bold=(r == 4), fill=GREY if r == 4 else None,
                             align="right")
        if val is not None:
            _amt(ws.cell(r, 3), val)
        elif r == 4:
            _f(ws.cell(r, 3), bold=True, fill=GREY)
        r += 1
    r += 1
    ws.cell(r, 1, "Total sale value — CG"); _f(ws.cell(r, 1), bold=True)
    _amt(ws.cell(r, 3), t["cg"], bold=True); r += 1
    ws.cell(r, 1, "Total sale value — AIS"); _f(ws.cell(r, 1), bold=True)
    _amt(ws.cell(r, 3), t["ais"], bold=True); r += 1
    ws.cell(r, 1, "Net delta (CG − AIS)"); _f(ws.cell(r, 1), bold=True)
    _amt(ws.cell(r, 3), t["delta"], bold=True, fill=YELLOW)
    for col, w in zip("ABC", (28, 12, 18)):
        ws.column_dimensions[col].width = w


def _pair_header(ws, cg_label, ais_label):
    hdr = ["Security", "ISIN", f"{cg_label} value", f"{ais_label} value",
           "Delta (CG−AIS)", "CG rows", "AIS rows"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h); _f(ws.cell(3, j), bold=True, fill=GREY)


def _mismatch_sheet(wb, result, cg_label, ais_label):
    ws = wb.create_sheet("Mismatched")
    _pair_header(ws, cg_label, ais_label)
    for i, p in enumerate(result.mismatched):
        row = 4 + i
        ws.cell(row, 1, p.name); _f(ws.cell(row, 1))
        ws.cell(row, 2, p.isin); _f(ws.cell(row, 2))
        _amt(ws.cell(row, 3), p.cg_value)
        _amt(ws.cell(row, 4), p.ais_value)
        _amt(ws.cell(row, 5), p.delta, fill=YELLOW)
        ws.cell(row, 6, p.cg_n); _f(ws.cell(row, 6), align="right")
        ws.cell(row, 7, p.ais_n); _f(ws.cell(row, 7), align="right")
    _widths(ws)
    ws.freeze_panes = "A4"


def _matched_sheet(wb, result, cg_label, ais_label):
    ws = wb.create_sheet("Matched")
    _pair_header(ws, cg_label, ais_label)
    for i, p in enumerate(result.matched):
        row = 4 + i
        ws.cell(row, 1, p.name); _f(ws.cell(row, 1))
        ws.cell(row, 2, p.isin); _f(ws.cell(row, 2))
        _amt(ws.cell(row, 3), p.cg_value)
        _amt(ws.cell(row, 4), p.ais_value)
        _amt(ws.cell(row, 5), p.delta)
        ws.cell(row, 6, p.cg_n); _f(ws.cell(row, 6), align="right")
        ws.cell(row, 7, p.ais_n); _f(ws.cell(row, 7), align="right")
    _widths(ws)
    ws.freeze_panes = "A4"


def _side_sheet(wb, title, sides, fill):
    ws = wb.create_sheet(title)
    hdr = ["Security", "ISIN", "Sale value", "Qty", "Rows"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h); _f(ws.cell(3, j), bold=True, fill=GREY)
    for i, s in enumerate(sides):
        row = 4 + i
        ws.cell(row, 1, s.name); _f(ws.cell(row, 1))
        ws.cell(row, 2, s.isin); _f(ws.cell(row, 2))
        _amt(ws.cell(row, 3), round(s.value, 2), fill=fill)
        _amt(ws.cell(row, 4), round(s.qty, 2))
        ws.cell(row, 5, s.n); _f(ws.cell(row, 5), align="right")
    for col, w in zip("ABCDE", (40, 16, 18, 14, 8)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def _widths(ws):
    for col, w in zip("ABCDEFG", (40, 16, 18, 18, 18, 10, 10)):
        ws.column_dimensions[col].width = w
