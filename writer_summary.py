"""
writer_summary.py — Output A: firm CG deliverable (xlsx, Arial 11).

Three sheets:
  - "CG Summary"      : six buckets + totals, the COI-feeding view
  - "Workings"        : every lot with the FULL audit snapshot (classification basis +
                        computation basis + grandfathering detail) — the "logic snapshot"
  - "COI block"       : paste-ready computation block, cross-referenced to CG Summary

Gains are live formulas (net - cost) so edits flow through. compute.py remains the
source of truth; the sheet ties to it on first open.
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
SZ = 11
ACCT = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
DATEFMT = "DD-MMM-YYYY"
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREY = PatternFill("solid", fgColor="EFEFEF")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BUCKET_ORDER = [
    "LTCG 112A Equity", "STCG 111A Equity", "LTCG Debt", "STCG Debt (50AA/slab)",
    "VDA 115BBH", "Equity non-STT (LTCG)", "Equity non-STT (STCG)",
    "Foreign (LTCG)", "Foreign (STCG)", "Unlisted (LTCG)", "Unlisted (STCG)",
]


def _f(cell, bold=False, size=SZ, fill=None, align=None, wrap=False):
    cell.font = Font(name=ARIAL, size=size, bold=bold)
    if fill: cell.fill = fill
    if align: cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    else: cell.alignment = Alignment(vertical="center", wrap_text=wrap)


def _amt(cell, formula_or_val, bold=False, fill=None):
    cell.value = formula_or_val
    cell.number_format = ACCT
    _f(cell, bold=bold, fill=fill)


def write_summary(results, path, client="", ay="2025-26", validation=None):
    wb = Workbook()
    _summary_sheet(wb, results, client, ay)
    _workings_sheet(wb, results, client, ay)
    _coi_sheet(wb, results, client, ay)
    if validation is not None:
        # fold the broker-vs-engine check into the deliverable (Output A is self-checking)
        from writer_validation import add_validation_sheets
        add_validation_sheets(wb, validation, client)
    wb.remove(wb["Sheet"]) if "Sheet" in wb.sheetnames else None
    wb.save(path)
    return path


def _by_bucket(results):
    d = {}
    for r in results:
        d.setdefault(r.bucket, []).append(r)
    return d


# ---- Summary sheet -------------------------------------------------------

def _summary_sheet(wb, results, client, ay):
    ws = wb.create_sheet("CG Summary")
    buckets = _by_bucket(results)
    ws["A1"] = client or "Capital Gain Summary"; _f(ws["A1"], bold=True, size=12)
    ws["A2"] = f"AY {ay}  |  pivot 23-Jul-2024  |  computer-prepared, preparer to verify"
    _f(ws["A2"], size=9)
    row = 4
    grand_sale, grand_cost, grand_gain = [], [], []
    n = 1
    for b in BUCKET_ORDER:
        rows = buckets.get(b)
        if not rows:
            continue
        ws.cell(row, 1, f"{n}.) {b}"); _f(ws.cell(row, 1), bold=True); row += 1
        hdr = ["Security", "ISIN", "Section / Rate", "Sale Consideration",
               "Cost of Acquisition", "Gain / (Loss)"]
        for j, h in enumerate(hdr, 2):
            ws.cell(row, j, h); _f(ws.cell(row, j), bold=True, fill=GREY)
        row += 1
        first = row
        for r in rows:
            ws.cell(row, 2, r.tx.security_name); _f(ws.cell(row, 2))
            ws.cell(row, 3, r.tx.isin or ""); _f(ws.cell(row, 3))
            ws.cell(row, 4, r.rate_label); _f(ws.cell(row, 4))
            _amt(ws.cell(row, 5), round(r.net_sale_consideration, 2))
            _amt(ws.cell(row, 6), round(r.cost_used, 2))
            _amt(ws.cell(row, 7), f"=E{row}-F{row}", fill=YELLOW)
            row += 1
        ws.cell(row, 2, "Total"); _f(ws.cell(row, 2), bold=True)
        for col in (5, 6, 7):
            L = get_column_letter(col)
            _amt(ws.cell(row, col), f"=SUM({L}{first}:{L}{row-1})", bold=True)
        grand_sale.append(f"E{row}"); grand_cost.append(f"F{row}"); grand_gain.append(f"G{row}")
        row += 2; n += 1

    ws.cell(row, 2, "GRAND TOTAL"); _f(ws.cell(row, 2), bold=True)
    if grand_sale:
        _amt(ws.cell(row, 5), "=" + "+".join(grand_sale), bold=True)
        _amt(ws.cell(row, 6), "=" + "+".join(grand_cost), bold=True)
        _amt(ws.cell(row, 7), "=" + "+".join(grand_gain), bold=True)
    for col, w in zip("ABCDEFG", (4, 42, 16, 18, 18, 18, 18)):
        ws.column_dimensions[col].width = w


# ---- Workings sheet (the logic snapshot) ---------------------------------

def _workings_sheet(wb, results, client, ay):
    ws = wb.create_sheet("Workings")
    ws["A1"] = f"{client} — CG Workings (classification & computation snapshot)"
    _f(ws["A1"], bold=True, size=12)
    hdr = ["Source", "Security", "ISIN", "Asset type", "Classification basis", "Conf.",
           "Acq date", "Transfer date", "Hold days", "Thr (m)", "LTCG?", "Pivot",
           "Section", "Rate", "Cost basis", "GF applied", "GF detail",
           "Qty", "Sale consid.", "Sell exp.", "Net sale", "Cost used", "Gain", "Flags"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h); _f(ws.cell(3, j), bold=True, fill=GREY, wrap=True)
    r0 = 4
    for i, r in enumerate(results):
        t = r.tx; row = r0 + i
        vals = [t.source_label, t.security_name, t.isin or "", t.asset_type,
                t.classification_basis, t.classification_confidence,
                t.acquisition_date, t.transfer_date, r.holding_days, r.threshold_months,
                "Yes" if r.is_ltcg else "No", r.pivot_side, r.section, r.rate_label,
                t.cost_basis_meaning, "Yes" if r.grandfathering_applied else "No",
                r.grandfathering_detail, t.quantity]
        for j, v in enumerate(vals, 1):
            ws.cell(row, j, v); _f(ws.cell(row, j), size=10, wrap=(j in (5, 17, 24)))
        ws.cell(row, 7).number_format = DATEFMT
        ws.cell(row, 8).number_format = DATEFMT
        _amt(ws.cell(row, 19), round(t.sale_consideration, 2))
        _amt(ws.cell(row, 20), round(t.transfer_expenses, 2))
        _amt(ws.cell(row, 21), round(r.net_sale_consideration, 2))
        _amt(ws.cell(row, 22), round(r.cost_used, 2))
        _amt(ws.cell(row, 23), f"=U{row}-V{row}", fill=YELLOW)
        ws.cell(row, 24, "; ".join(r.flags)); _f(ws.cell(row, 24), size=9, wrap=True)
    widths = [12, 34, 14, 12, 26, 8, 12, 12, 8, 7, 6, 6, 12, 16, 12, 8, 34, 10, 15, 12, 15, 15, 15, 28]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A4"


# ---- COI block -----------------------------------------------------------

_COI = [
    ("Long Term Capital Gain 112A: (Equity)", "LTCG 112A Equity"),
    ("Long Term Capital Gain: (Debt)", "LTCG Debt"),
    ("Short Term Capital Gain 111A: (Equity)", "STCG 111A Equity"),
    ("Short Term Capital Gain: (Debt)", "STCG Debt (50AA/slab)"),
    ("Income from VDA (115BBH)", "VDA 115BBH"),
    ("Sale of Foreign Shares (LTCG)", "Foreign (LTCG)"),
    ("Sale of Foreign Shares (STCG)", "Foreign (STCG)"),
    ("Sale of Unlisted Shares (LTCG)", "Unlisted (LTCG)"),
    ("Sale of Unlisted Shares (STCG)", "Unlisted (STCG)"),
]


def _coi_sheet(wb, results, client, ay):
    ws = wb.create_sheet("COI block")
    buckets = _by_bucket(results)
    ws["B1"] = "INCOME FROM CAPITAL GAINS (Refer CG workings)"; _f(ws["B1"], bold=True, size=12)
    row = 3
    for label, bkey in _COI:
        if not buckets.get(bkey):
            continue
        ws.cell(row, 2, label); _f(ws.cell(row, 2), bold=True); row += 1
        ws.cell(row, 2, "Sale Consideration"); _f(ws.cell(row, 2))
        ws.cell(row, 4, "Rs."); _f(ws.cell(row, 4))
        sale = sum(round(r.net_sale_consideration, 2) for r in buckets[bkey])
        cost = sum(round(r.cost_used, 2) for r in buckets[bkey])
        _amt(ws.cell(row, 5), sale); srow = row; row += 1
        ws.cell(row, 2, "Less: Cost of Acquisition"); _f(ws.cell(row, 2))
        ws.cell(row, 4, "Rs."); _amt(ws.cell(row, 5), cost); crow = row; row += 1
        ws.cell(row, 2, "Capital Gain"); _f(ws.cell(row, 2), bold=True)
        ws.cell(row, 4, "Rs."); _amt(ws.cell(row, 5), f"=E{srow}-E{crow}", bold=True); row += 1
        ws.cell(row, 2, "Less: PY Loss C/F set off"); _f(ws.cell(row, 2))
        ws.cell(row, 4, "Rs."); ws.cell(row, 5).fill = YELLOW
        ws.cell(row, 5).comment = None; _f(ws.cell(row, 5)); pyrow = row; row += 1
        ws.cell(row, 2, "Balance Gain / (Loss)"); _f(ws.cell(row, 2), bold=True)
        ws.cell(row, 4, "Rs."); _amt(ws.cell(row, 5), f"=E{srow}-E{crow}-E{pyrow}", bold=True)
        row += 2
    for col, w in zip("ABCDE", (3, 40, 4, 6, 18)):
        ws.column_dimensions[col].width = w
