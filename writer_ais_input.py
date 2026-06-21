"""
writer_ais_input.py — Output D: AIS-reconciliation input (xlsx, Arial 11).

The CG engine's clean SALE side, shaped so it drops straight into the AIS
Reconciliation path as the "capital-gains / broker file". Re-uploading the raw
broker statement there means re-detecting its messy columns and verbose names;
feeding THIS file instead hands the reconciliation a pristine, ISIN-keyed sale
side, so every lot matches its AIS line by ISIN with nothing lost to a name
mismatch. The headers are the canonical ones the detector knows, so the AIS path
auto-maps it without a single override.

Only the securities AIS's "Sale of securities and units of mutual fund" reports
are included — listed equity, equity / other MF units, business trusts and
(non-equity) MF / debt units, all reported to AIS by the depositories and RTAs.
Foreign and unlisted shares are not in that AIS section, and VDA sales are
reported in their own AIS section, so those are set aside on a separate sheet
(with the reason) rather than polluting the reconciliation as spurious
"only in CG" rows. Nothing is dropped silently.

Sheets:
  - "AIS Reco Input"      : lot-level — ISIN, name, sale date, qty, gross sale value.
  - "By ISIN (TIS view)"  : the same totalled per security (ISIN), to eyeball directly
                            against the AIS / TIS per-ISIN sale figures.
  - "Not in AIS securities": foreign / unlisted / VDA set aside, with the reason.

Mirrors the other writers' look (Arial 11, accounting format, grey headers).
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import reco  # reuse the exact per-security keying the reconciliation will later apply

ARIAL = "Arial"
SZ = 11
ACCT = "#,##0.00"
QTY = "#,##0.000"
DATEFMT = "DD-MMM-YYYY"
GREY = PatternFill("solid", fgColor="EFEFEF")

# The asset classes AIS's "Sale of securities and units of mutual fund" covers.
AIS_ASSETS = {"equity", "eof", "business_trust", "mf_debt"}

ASSET_LABEL = {
    "equity": "Listed equity share",
    "eof": "Equity / other MF unit",
    "business_trust": "Business trust (REIT / InvIT)",
    "mf_debt": "MF unit (non-equity) / debt",
    "foreign": "Foreign security",
    "unlisted": "Unlisted share",
    "vda": "Virtual digital asset",
}

# Why a set-aside row is not in the AIS securities reconciliation.
WHY_EXCLUDED = {
    "foreign": "foreign — not reported in the Indian AIS securities section",
    "unlisted": "unlisted — not in the AIS securities section",
    "vda": "VDA — reported in the AIS VDA section; reconcile separately",
}


def _hdr(cell, text):
    cell.value = text
    cell.font = Font(name=ARIAL, size=SZ, bold=True)
    cell.fill = GREY
    cell.alignment = Alignment(vertical="center")


def _cell(cell, val=None, num=None, align=None, bold=False):
    if val is not None:
        cell.value = val
    cell.font = Font(name=ARIAL, size=SZ, bold=bold)
    if num:
        cell.number_format = num
    cell.alignment = Alignment(horizontal=align, vertical="center")


def write_ais_input(results, path, client="Client"):
    wb = Workbook()
    wb.remove(wb.active)
    incl = [r for r in results if r.tx.asset_type in AIS_ASSETS]
    excl = [r for r in results if r.tx.asset_type not in AIS_ASSETS]
    _lot_sheet(wb, incl)
    _by_isin_sheet(wb, incl)
    if excl:
        _excluded_sheet(wb, excl)
    wb.save(path)
    return path


def _lot_sheet(wb, rows):
    """Lot-level sale side with canonical headers — the machine feed for the AIS path.
    Sale value is the GROSS consideration (what the depository reports to AIS), not net
    of the broker's charges, so it lines up with the AIS figure."""
    ws = wb.create_sheet("AIS Reco Input")
    headers = ["ISIN", "Security Name", "Asset Class", "Sale Date",
               "Quantity", "Sale Consideration"]
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(1, j), h)
    for i, r in enumerate(rows):
        t = r.tx
        row = 2 + i
        _cell(ws.cell(row, 1), t.isin or "")
        _cell(ws.cell(row, 2), t.security_name or "")
        _cell(ws.cell(row, 3), ASSET_LABEL.get(t.asset_type, t.asset_type))
        if t.transfer_date is not None:
            _cell(ws.cell(row, 4), t.transfer_date, num=DATEFMT)
        else:
            _cell(ws.cell(row, 4))
        _cell(ws.cell(row, 5),
              t.quantity if t.quantity is not None else "", num=QTY, align="right")
        _cell(ws.cell(row, 6), round(t.sale_consideration, 2), num=ACCT, align="right")
    for col, w in zip("ABCDEF", (16, 40, 26, 14, 14, 18)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def _by_isin_sheet(wb, rows):
    """The same sale side totalled per security, keyed exactly as the reconciliation
    will key it (ISIN first), so this is the CG side of the AIS/TIS comparison the
    preparer can eyeball against AIS's per-ISIN totals before even running the reco."""
    ws = wb.create_sheet("By ISIN (TIS view)")
    headers = ["ISIN", "Security Name", "Asset Class", "Lots",
               "Total Quantity", "Total Sale Consideration"]
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(1, j), h)
    agg: dict = {}
    for r in rows:
        t = r.tx
        key, kind = reco.reco_key(t.isin, t.security_name)
        if not key:
            continue
        a = agg.get(key)
        if a is None:
            agg[key] = {"isin": key if kind == "isin" else "",
                        "name": t.security_name or "",
                        "asset": ASSET_LABEL.get(t.asset_type, t.asset_type),
                        "qty": t.quantity or 0.0,
                        "val": t.sale_consideration or 0.0, "n": 1}
        else:
            a["qty"] += t.quantity or 0.0
            a["val"] += t.sale_consideration or 0.0
            a["n"] += 1
            if not a["name"] and t.security_name:
                a["name"] = t.security_name
    ordered = sorted(agg.values(), key=lambda a: -abs(a["val"]))
    for i, a in enumerate(ordered):
        row = 2 + i
        _cell(ws.cell(row, 1), a["isin"])
        _cell(ws.cell(row, 2), a["name"])
        _cell(ws.cell(row, 3), a["asset"])
        _cell(ws.cell(row, 4), a["n"], align="right")
        _cell(ws.cell(row, 5), round(a["qty"], 3), num=QTY, align="right")
        _cell(ws.cell(row, 6), round(a["val"], 2), num=ACCT, align="right")
    for col, w in zip("ABCDEF", (16, 40, 26, 8, 16, 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def _excluded_sheet(wb, rows):
    ws = wb.create_sheet("Not in AIS securities")
    ws["A1"] = ("Set aside from the AIS securities reconciliation — reported in a "
                "different AIS section, or not by the Indian depositories.")
    _cell(ws["A1"], bold=True)
    headers = ["Security", "ISIN", "Asset Class", "Section",
               "Sale Consideration", "Why set aside"]
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(3, j), h)
    for i, r in enumerate(rows):
        t = r.tx
        row = 4 + i
        _cell(ws.cell(row, 1), t.security_name or "")
        _cell(ws.cell(row, 2), t.isin or "")
        _cell(ws.cell(row, 3), ASSET_LABEL.get(t.asset_type, t.asset_type))
        _cell(ws.cell(row, 4), r.section)
        _cell(ws.cell(row, 5), round(t.sale_consideration, 2), num=ACCT, align="right")
        _cell(ws.cell(row, 6), WHY_EXCLUDED.get(t.asset_type, "set aside"))
    for col, w in zip("ABCDEF", (34, 16, 26, 18, 18, 46)):
        ws.column_dimensions[col].width = w
