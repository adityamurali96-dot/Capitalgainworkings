"""
writer_winman.py — Output B: Winman-import-shaped xlsx (Arial 11).

Plain xlsx (no macros) carrying ONLY the input columns Winman needs, in Winman's
column order, computed/grey columns present-but-blank. You paste the data block
into your live `securitiesshortGain.xlsm` (or import if your Winman build accepts
xlsx). The macro computes NETSALE / COSTOFACQUISITION / SHORTTERM / LTCG itself —
we never pre-fill those, so there is one source of truth for the gain.

Three data sheets, routed by asset_type:
  - "Gains on STT paid shares"        <- equity, eof, business_trust  (Sheet 1)
  - "Units of MF except Equity fund"  <- mf_debt                      (Sheet 3)
  - "Virtual Digital Assets"          <- vda                          (Sheet 5)
foreign/unlisted/non-STT equity are NOT in this template — they are listed on a
"Not in Winman" sheet for the ITR schedules instead.

NOTE: column order below mirrors the documented machine-key row. Eyeball the
first paste against your Winman version's columns and tweak ORDER_* if a build
differs — this is the one spot that is build-specific.
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ARIAL = "Arial"; SZ = 11
GREY = PatternFill("solid", fgColor="D9D9D9")
TAG = PatternFill("solid", fgColor="FFF2CC")
DATEFMT = "DD-MMM-YYYY"
ACCT = "#,##0.00"

# (machine_key, human_label, source_field_or_None_if_computed)
SHEET1 = [
    ("PARTICULARS", "Particulars", "security_name"),
    ("QUANTITY", "Quantity", "quantity"),
    ("PURCHASEDATE", "Date of Purchase", "acquisition_date"),
    ("TRANSFERDATE", "Date of Transfer", "transfer_date"),
    ("ISITLTCG", "Is it LTCG?", "_is_ltcg"),
    ("SALE", "Sale consideration", "sale_consideration"),
    ("SELLING", "Selling Expenses", "transfer_expenses"),
    ("NETSALE", "Net Sale consideration", None),          # grey
    ("ACTUALCOST", "Actual cost of Acquisition", "_actualcost"),
    ("ACQEXPENSES", "Other acq. expenses (incl. in cost)", "purchase_expenses"),
    ("FMV", "FMV u/s 55(2)(ac) per share/unit", "fmv_31jan2018"),
    ("COSTOFACQUISITION", "Cost of Acquisition deductible", None),  # grey
    ("SHORTTERM", "Short term gain u/s 111A", None),      # grey
    ("LTCG", "LTCG u/s 112A", None),                      # grey
    ("ISIN", "ISIN code", "isin"),
]
SHEET3 = [
    ("PARTICULARS", "Particulars", "security_name"),
    ("TYPE", "Type", "_mf_type"),
    ("QUANTITY", "Quantity", "quantity"),
    ("PURCHASEDATE", "Date of Purchase", "acquisition_date"),
    ("TRANSFERDATE", "Date of Transfer", "transfer_date"),
    ("ISITLTCG", "Is it LTCG?", "_is_ltcg"),
    ("SALE", "Sale consideration", "sale_consideration"),
    ("SELLING", "Selling Expenses", "transfer_expenses"),
    ("NETSALE", "Net Sale consideration", None),
    ("ACTUALCOST", "Actual cost of Acquisition", "_actualcost"),
    ("ACQEXPENSES", "Other acq. expenses (incl. in cost)", "purchase_expenses"),
    ("INDEXED", "Indexed cost (if applicable)", None),
    ("SHORTTERM", "Short term gain", None),
    ("LTCG", "LTCG", None),
    ("ISIN", "ISIN code", "isin"),
]
SHEET5 = [
    ("PARTICULARS", "Particulars", "security_name"),
    ("QUANTITY", "Quantity", "quantity"),
    ("PURCHASEDATE", "Date of Acquisition", "acquisition_date"),
    ("TRANSFERDATE", "Date of Transfer", "transfer_date"),
    ("SALE", "Sale consideration", "sale_consideration"),
    ("ACTUALCOST", "Cost of Acquisition", "_actualcost"),
    ("ACQEXPENSES", "Other acq. expenses (incl. in cost)", "purchase_expenses"),
    ("INCOME", "Income u/s 115BBH", None),
    ("ISIN", "ISIN / token", "isin"),
]

ROUTE = {
    "equity": ("Gains on STT paid shares", SHEET1),
    "eof": ("Gains on STT paid shares", SHEET1),
    "business_trust": ("Gains on STT paid shares", SHEET1),
    "mf_debt": ("Units of MF except Equity fund", SHEET3),
    "vda": ("Virtual Digital Assets", SHEET5),
}


def _val(r, field):
    t = r.tx
    if field == "_is_ltcg":
        return "Yes" if r.is_ltcg else "No"
    if field == "_actualcost":
        # Actual cost of acquisition = stated purchase cost PLUS any mapped other
        # expenses on acquisition (deductible — the preparer opted in by mapping the
        # column). Winman does the grandfathering substitution from this figure; the
        # raw expense portion is also shown separately in the ACQEXPENSES column.
        return round(t.purchase_cost + (t.purchase_expenses or 0.0), 2)
    if field == "_mf_type":
        return "50AA - Debt oriented" if t.is_50aa else "Listed - others"
    if field is None:
        return None
    v = getattr(t, field, None)
    return v


def write_winman(results, path):
    wb = Workbook()
    wb.remove(wb.active)
    counts = {}
    for sheet_name, schema in [("Gains on STT paid shares", SHEET1),
                               ("Units of MF except Equity fund", SHEET3),
                               ("Virtual Digital Assets", SHEET5)]:
        ws = wb.create_sheet(sheet_name)
        rows = [r for r in results if ROUTE.get(r.tx.asset_type, (None,))[0] == sheet_name]
        counts[sheet_name] = len(rows)
        _write_sheet(ws, schema, rows)
    _not_in_winman(wb, results)
    wb.save(path)
    return path, counts


def _write_sheet(ws, schema, rows):
    # Row 1: machine-key tag row
    ws.cell(1, 1, "securitiesshortGain"); ws.cell(1, 1).fill = TAG
    for j, (mk, _, _) in enumerate(schema, 1):
        c = ws.cell(2, j, mk); c.font = Font(name=ARIAL, size=9, bold=True); c.fill = TAG
    # Row 3: human labels
    for j, (mk, label, src) in enumerate(schema, 1):
        c = ws.cell(3, j, label)
        c.font = Font(name=ARIAL, size=SZ, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if src is None:
            c.fill = GREY  # mark computed/grey columns
    # Row 4: dashed marker (data begins below)
    for j in range(1, len(schema) + 1):
        ws.cell(4, j, "-").font = Font(name=ARIAL, size=SZ)
    # Data from row 5
    for i, r in enumerate(rows):
        row = 5 + i
        for j, (mk, label, src) in enumerate(schema, 1):
            v = _val(r, src)
            c = ws.cell(row, j, v)
            c.font = Font(name=ARIAL, size=SZ)
            if src in ("acquisition_date", "transfer_date"):
                c.number_format = DATEFMT
            elif src in ("sale_consideration", "transfer_expenses", "purchase_expenses",
                         "_actualcost", "fmv_31jan2018"):
                c.number_format = ACCT
            if src is None:
                c.fill = GREY
    for j, (mk, label, src) in enumerate(schema, 1):
        ws.column_dimensions[get_column_letter(j)].width = 16 if src else 14
    ws.column_dimensions["A"].width = 36


def _not_in_winman(wb, results):
    out = [r for r in results if r.tx.asset_type in ("foreign", "unlisted")
           or (r.tx.asset_type in ("equity", "eof", "business_trust") and not r.tx.stt_paid)]
    if not out:
        return
    ws = wb.create_sheet("Not in Winman")
    ws["A1"] = "Outside the Winman CG template — enter in ITR schedules manually"
    ws["A1"].font = Font(name=ARIAL, size=SZ, bold=True)
    hdr = ["Security", "ISIN", "Asset type", "Section", "Sale consid.", "Cost used", "Gain", "Why"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h).font = Font(name=ARIAL, size=SZ, bold=True)
    for i, r in enumerate(out):
        row = 4 + i
        vals = [r.tx.security_name, r.tx.isin or "", r.tx.asset_type, r.section,
                round(r.net_sale_consideration, 2), round(r.cost_used, 2),
                round(r.gain, 2), "; ".join(r.flags)]
        for j, v in enumerate(vals, 1):
            ws.cell(row, j, v).font = Font(name=ARIAL, size=SZ)
    for col, w in zip("ABCDEFGH", (34, 14, 12, 16, 16, 16, 16, 40)):
        ws.column_dimensions[col].width = w
