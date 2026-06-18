"""
writer_validation.py — Output C: broker-vs-engine validation (xlsx, Arial 11).

Two sheets (also appended into Output A so the deliverable is self-checking):
  - "Validation"      : the short / long / total roll-up (engine vs broker),
                        the per-bucket breakdown, and the figures the broker
                        already PRINTED in the statement (for the eyeball check).
  - "Lot Validation"  : every lot, engine gain vs the broker's own stated gain,
                        with the delta and a status — mismatches highlighted.

Mirrors writer_summary / writer_reco (Arial 11, accounting format, thin borders).
Compares only; the preparer judges every delta.
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
SZ = 11
ACCT = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
GREY = PatternFill("solid", fgColor="EFEFEF")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREEN = PatternFill("solid", fgColor="E6F4EA")
RED = PatternFill("solid", fgColor="FBE6E6")
AMBER = PatternFill("solid", fgColor="FCF3D9")

_STATUS_FILL = {"match": GREEN, "mismatch": RED, "no_broker": GREY}
_STATUS_TEXT = {"match": "✓ match", "mismatch": "✗ differs", "no_broker": "— no broker figure"}


def _f(cell, bold=False, size=SZ, fill=None, align=None, wrap=False):
    cell.font = Font(name=ARIAL, size=size, bold=bold)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def _amt(cell, val, bold=False, fill=None):
    cell.value = val
    cell.number_format = ACCT
    _f(cell, bold=bold, fill=fill)


# ---- Validation summary sheet -----------------------------------------------

def _validation_sheet(wb, vres, client):
    ws = wb.create_sheet("Validation")
    ws["A1"] = f"{client or 'Capital Gain'} — Validation (broker vs engine)"
    _f(ws["A1"], bold=True, size=12)
    cov = vres.coverage
    ws["A2"] = (f"Broker per-lot gain present on {cov.get('n_broker', 0)} of "
                f"{cov.get('n', 0)} lots   |   computer-prepared, preparer to verify")
    _f(ws["A2"], size=9)

    row = 4
    if not vres.has_broker_gain():
        ws.cell(row, 1, "No broker per-lot gain column was mapped — map the broker's own "
                        "gain / P&L column on the Map screen to enable per-lot validation.")
        _f(ws.cell(row, 1), bold=True, fill=AMBER); row += 2

    # roll-up: short / long / total
    ws.cell(row, 1, "Roll-up"); _f(ws.cell(row, 1), bold=True); row += 1
    hdr = ["", "Engine gain", "Broker gain", "Delta (eng−brk)", "Status", "Lots (broker/total)"]
    for j, h in enumerate(hdr, 1):
        ws.cell(row, j, h); _f(ws.cell(row, j), bold=True, fill=GREY)
    row += 1
    for key, label in (("short", "Short term"), ("long", "Long term"), ("total", "TOTAL")):
        d = vres.rollup.get(key, {})
        bold = key == "total"
        ws.cell(row, 1, label); _f(ws.cell(row, 1), bold=bold)
        _amt(ws.cell(row, 2), d.get("engine", 0.0), bold=bold)
        if d.get("n_broker"):
            _amt(ws.cell(row, 3), d.get("broker", 0.0), bold=bold)
            _amt(ws.cell(row, 4), d.get("delta", 0.0), bold=bold, fill=YELLOW)
        else:
            ws.cell(row, 3, "—"); _f(ws.cell(row, 3))
            ws.cell(row, 4, "—"); _f(ws.cell(row, 4))
        st = d.get("status", "no_broker")
        ws.cell(row, 5, _STATUS_TEXT[st]); _f(ws.cell(row, 5), bold=bold, fill=_STATUS_FILL[st])
        ws.cell(row, 6, f"{d.get('n_broker', 0)} / {d.get('n', 0)}")
        _f(ws.cell(row, 6), align="right")
        row += 1
    row += 1

    # per-bucket breakdown
    ws.cell(row, 1, "By bucket"); _f(ws.cell(row, 1), bold=True); row += 1
    for j, h in enumerate(["Bucket", "Engine gain", "Broker gain", "Delta (eng−brk)",
                           "Status", "Lots (broker/total)"], 1):
        ws.cell(row, j, h); _f(ws.cell(row, j), bold=True, fill=GREY)
    row += 1
    for c in vres.buckets:
        ws.cell(row, 1, c.key); _f(ws.cell(row, 1))
        _amt(ws.cell(row, 2), c.engine_gain)
        if c.n_broker:
            _amt(ws.cell(row, 3), c.broker_gain)
            _amt(ws.cell(row, 4), c.delta, fill=(YELLOW if c.status == "mismatch" else None))
        else:
            ws.cell(row, 3, "—"); _f(ws.cell(row, 3))
            ws.cell(row, 4, "—"); _f(ws.cell(row, 4))
        ws.cell(row, 5, _STATUS_TEXT[c.status]); _f(ws.cell(row, 5), fill=_STATUS_FILL[c.status])
        ws.cell(row, 6, f"{c.n_broker} / {c.n}"); _f(ws.cell(row, 6), align="right")
        row += 1
    row += 1

    # broker's PRINTED figures (scanned from the statement, for the eyeball check)
    ws.cell(row, 1, "Figures the broker already printed (for reference)")
    _f(ws.cell(row, 1), bold=True); row += 1
    if vres.printed:
        for j, h in enumerate(["Sheet", "Label", "Type", "Value"], 1):
            ws.cell(row, j, h); _f(ws.cell(row, j), bold=True, fill=GREY)
        row += 1
        for p in vres.printed:
            ws.cell(row, 1, p.sheet); _f(ws.cell(row, 1))
            ws.cell(row, 2, p.label); _f(ws.cell(row, 2))
            ws.cell(row, 3, {"short": "Short term", "long": "Long term",
                             "total": "Total"}.get(p.kind, p.kind)); _f(ws.cell(row, 3))
            _amt(ws.cell(row, 4), p.value)
            row += 1
    else:
        ws.cell(row, 1, "No printed short/long-term gain figures were found in the workbook.")
        _f(ws.cell(row, 1), size=10); row += 1

    for col, w in zip("ABCDEF", (30, 44, 18, 18, 20, 18)):
        ws.column_dimensions[col].width = w


# ---- Lot-level validation sheet ---------------------------------------------

def _lot_sheet(wb, vres):
    ws = wb.create_sheet("Lot Validation")
    hdr = ["Security", "ISIN", "Section", "LTCG?", "Engine gain", "Broker gain",
           "Delta (eng−brk)", "Status"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h); _f(ws.cell(3, j), bold=True, fill=GREY)
    # mismatches first, then no-broker, then matches — the chase list on top
    order = {"mismatch": 0, "no_broker": 1, "match": 2}
    lots = sorted(vres.lots, key=lambda l: (order.get(l.status, 9), -abs(l.delta or 0)))
    for i, l in enumerate(lots):
        row = 4 + i
        ws.cell(row, 1, l.security_name); _f(ws.cell(row, 1))
        ws.cell(row, 2, l.isin); _f(ws.cell(row, 2))
        ws.cell(row, 3, l.section); _f(ws.cell(row, 3))
        ws.cell(row, 4, "Yes" if l.is_ltcg else "No"); _f(ws.cell(row, 4))
        _amt(ws.cell(row, 5), l.engine_gain)
        if l.broker_gain is not None:
            _amt(ws.cell(row, 6), l.broker_gain)
            _amt(ws.cell(row, 7), l.delta, fill=(YELLOW if l.status == "mismatch" else None))
        else:
            ws.cell(row, 6, "—"); _f(ws.cell(row, 6))
            ws.cell(row, 7, "—"); _f(ws.cell(row, 7))
        ws.cell(row, 8, _STATUS_TEXT[l.status]); _f(ws.cell(row, 8), fill=_STATUS_FILL[l.status])
    for col, w in zip("ABCDEFGH", (40, 16, 20, 8, 16, 16, 18, 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# ---- public API -------------------------------------------------------------

def add_validation_sheets(wb, vres, client=""):
    """Append the Validation + Lot Validation sheets to an existing workbook
    (used to fold validation into Output A — the CG Summary file)."""
    _validation_sheet(wb, vres, client)
    _lot_sheet(wb, vres)
    return wb


def write_validation(vres, path, client=""):
    """Standalone Output C workbook."""
    wb = Workbook()
    add_validation_sheets(wb, vres, client)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(path)
    return path
